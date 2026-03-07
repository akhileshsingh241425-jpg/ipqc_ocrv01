#!/usr/bin/env python3
"""
IQC (Incoming Quality Control) Excel Report Generator
Generates formatted IQC inspection report for each raw material.
Uses openpyxl for Excel generation with professional formatting.
"""

import sys
import json
import os
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ========== MATERIAL DEFINITIONS ==========
MATERIAL_COLORS = {
    'glass': '3B82F6',
    'eva': '8B5CF6',
    'cell': 'F59E0B',
    'backsheet': '06B6D4',
    'ribbon': '10B981',
    'frame': '64748B',
    'jbox': 'EF4444',
    'silicone': 'A855F7',
    'flux': 'EC4899',
    'tpt': '14B8A6',
    'label': '78716C',
}

MATERIAL_NAMES = {
    'glass': 'Glass',
    'eva': 'EVA (Ethylene Vinyl Acetate)',
    'cell': 'Solar Cells',
    'backsheet': 'Backsheet',
    'ribbon': 'Ribbon (Tabbing/Stringing)',
    'frame': 'Aluminium Frame',
    'jbox': 'Junction Box',
    'silicone': 'Silicone Sealant',
    'flux': 'Soldering Flux',
    'tpt': 'TPT Backsheet',
    'label': 'Label / Sticker',
}


def generate_iqc_excel(record, output_path):
    """Generate a formatted IQC inspection Excel report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'IQC Inspection'
    
    mat_type = record.get('materialType', 'unknown')
    mat_name = record.get('materialName', MATERIAL_NAMES.get(mat_type, mat_type))
    color_hex = MATERIAL_COLORS.get(mat_type, '1E293B')
    
    # Styles
    header_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    sub_header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    sub_header_font = Font(name='Calibri', size=11, bold=True, color='1E293B')
    cell_font = Font(name='Calibri', size=11, color='334155')
    label_font = Font(name='Calibri', size=10, bold=True, color='475569')
    ok_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
    ok_font = Font(name='Calibri', size=11, bold=True, color='059669')
    ng_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    ng_font = Font(name='Calibri', size=11, bold=True, color='DC2626')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 20
    
    row = 1
    
    # ===== TITLE =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title_cell = ws.cell(row=row, column=1, value=f'IQC INSPECTION REPORT — {mat_name.upper()}')
    title_cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    title_cell.fill = header_fill
    title_cell.alignment = center_align
    ws.row_dimensions[row].height = 40
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = header_fill
        ws.cell(row=row, column=c).border = thin_border
    row += 1
    
    # ===== SUBTITLE =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    sub_cell = ws.cell(row=row, column=1, value='GAUTAM SOLAR PVT. LTD. — Incoming Quality Control')
    sub_cell.font = Font(name='Calibri', size=11, italic=True, color='FFFFFF')
    sub_cell.fill = PatternFill(start_color='475569', end_color='475569', fill_type='solid')
    sub_cell.alignment = center_align
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = PatternFill(start_color='475569', end_color='475569', fill_type='solid')
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 28
    row += 2
    
    # ===== GENERAL INFO =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    sec = ws.cell(row=row, column=1, value='📝  GENERAL INFORMATION')
    sec.font = sub_header_font
    sec.fill = sub_header_fill
    sec.alignment = left_align
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = sub_header_fill
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 28
    row += 1
    
    general_fields = [
        ('Inspection Date', record.get('inspectionDate', '')),
        ('Inspector Name', record.get('inspectorName', '')),
        ('Supplier Name', record.get('supplierName', '')),
        ('Batch / Lot No', record.get('batchNo', '')),
        ('PO Number', record.get('poNumber', '')),
        ('Challan No', record.get('challanNo', '')),
        ('Qty Received', record.get('qtyReceived', '')),
        ('Qty Accepted', record.get('qtyAccepted', '')),
        ('Qty Rejected', record.get('qtyRejected', '')),
    ]
    
    # Write 3 per row (label, value pairs across 6 columns)
    for i in range(0, len(general_fields), 3):
        chunk = general_fields[i:i+3]
        for j, (lbl, val) in enumerate(chunk):
            col_lbl = 1 + j * 2
            col_val = 2 + j * 2
            lbl_cell = ws.cell(row=row, column=col_lbl, value=lbl)
            lbl_cell.font = label_font
            lbl_cell.alignment = left_align
            lbl_cell.border = thin_border
            val_cell = ws.cell(row=row, column=col_val, value=val)
            val_cell.font = cell_font
            val_cell.alignment = left_align
            val_cell.border = thin_border
        ws.row_dimensions[row].height = 24
        row += 1
    
    row += 1
    
    # ===== PARAMETER INSPECTION =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    sec2 = ws.cell(row=row, column=1, value='🔬  PARAMETER-WISE INSPECTION')
    sec2.font = sub_header_font
    sec2.fill = sub_header_fill
    sec2.alignment = left_align
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = sub_header_fill
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 28
    row += 1
    
    # Table header
    headers = ['#', 'Parameter', 'Observed Value', 'Result', 'Specification', 'Remarks']
    header_fill_tbl = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    for col_idx, h in enumerate(headers, 1):
        hc = ws.cell(row=row, column=col_idx, value=h)
        hc.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        hc.fill = header_fill_tbl
        hc.alignment = center_align
        hc.border = thin_border
    ws.row_dimensions[row].height = 28
    row += 1
    
    # Parameter rows
    params = record.get('params', {})
    param_results = record.get('paramResults', {})
    param_keys = list(params.keys()) if params else []
    
    ok_count = 0
    ng_count = 0
    
    for idx, param in enumerate(param_keys, 1):
        value = params.get(param, '')
        result = param_results.get(param, 'OK')
        
        ws.cell(row=row, column=1, value=idx).font = cell_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value=param).font = Font(name='Calibri', size=11, bold=True, color='334155')
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3, value=value).font = cell_font
        ws.cell(row=row, column=3).alignment = center_align
        ws.cell(row=row, column=3).border = thin_border
        
        result_cell = ws.cell(row=row, column=4, value=result)
        if result == 'OK':
            result_cell.font = ok_font
            result_cell.fill = ok_fill
            ok_count += 1
        else:
            result_cell.font = ng_font
            result_cell.fill = ng_fill
            ng_count += 1
        result_cell.alignment = center_align
        result_cell.border = thin_border
        
        ws.cell(row=row, column=5, value='—').font = cell_font
        ws.cell(row=row, column=5).alignment = center_align
        ws.cell(row=row, column=5).border = thin_border
        
        ws.cell(row=row, column=6, value='').font = cell_font
        ws.cell(row=row, column=6).alignment = left_align
        ws.cell(row=row, column=6).border = thin_border
        
        ws.row_dimensions[row].height = 24
        row += 1
    
    row += 1
    
    # ===== OVERALL RESULT =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    sec3 = ws.cell(row=row, column=1, value='📋  OVERALL RESULT')
    sec3.font = sub_header_font
    sec3.fill = sub_header_fill
    sec3.alignment = left_align
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = sub_header_fill
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 28
    row += 1
    
    overall = record.get('overallResult', 'Accepted')
    remarks = record.get('remarks', '')
    
    # Summary row
    summary_items = [
        ('OK Parameters', str(ok_count)),
        ('NG Parameters', str(ng_count)),
        ('Total Parameters', str(len(param_keys))),
        ('Overall Verdict', overall),
        ('', ''),
        ('', ''),
    ]
    for j, (lbl, val) in enumerate(summary_items[:3]):
        col_lbl = 1 + j * 2
        col_val = 2 + j * 2
        ws.cell(row=row, column=col_lbl, value=lbl).font = label_font
        ws.cell(row=row, column=col_lbl).alignment = left_align
        ws.cell(row=row, column=col_lbl).border = thin_border
        ws.cell(row=row, column=col_val, value=val).font = Font(name='Calibri', size=12, bold=True, color='1E293B')
        ws.cell(row=row, column=col_val).alignment = center_align
        ws.cell(row=row, column=col_val).border = thin_border
    ws.row_dimensions[row].height = 28
    row += 1
    
    # Overall verdict with color
    ws.cell(row=row, column=1, value='Overall Verdict').font = label_font
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=1).border = thin_border
    
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    verdict_cell = ws.cell(row=row, column=2, value=overall)
    if overall == 'Accepted':
        verdict_cell.font = Font(name='Calibri', size=14, bold=True, color='059669')
        verdict_cell.fill = ok_fill
    elif overall == 'Rejected':
        verdict_cell.font = Font(name='Calibri', size=14, bold=True, color='DC2626')
        verdict_cell.fill = ng_fill
    else:
        verdict_cell.font = Font(name='Calibri', size=14, bold=True, color='CA8A04')
        verdict_cell.fill = PatternFill(start_color='FEFCE8', end_color='FEFCE8', fill_type='solid')
    verdict_cell.alignment = center_align
    verdict_cell.border = thin_border
    for c in range(2, 7):
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 32
    row += 1
    
    # Remarks
    if remarks:
        ws.cell(row=row, column=1, value='Remarks').font = label_font
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=1).border = thin_border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2, value=remarks).font = cell_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        ws.row_dimensions[row].height = 40
        row += 1
    
    row += 2
    
    # ===== SIGNATURES =====
    ws.cell(row=row, column=1, value='Inspector Signature:').font = label_font
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=2, value='________________').font = cell_font
    
    ws.cell(row=row, column=4, value='QC Head Signature:').font = label_font
    ws.cell(row=row, column=4).alignment = left_align
    ws.cell(row=row, column=5, value='________________').font = cell_font
    row += 2
    
    ws.cell(row=row, column=1, value=f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}').font = Font(name='Calibri', size=9, italic=True, color='94A3B8')
    
    # Page setup
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.print_area = f'A1:F{row}'
    
    wb.save(output_path)
    print(f'IQC Excel saved: {output_path}')
    return True


def main():
    parser = argparse.ArgumentParser(description='IQC Excel Report Generator')
    parser.add_argument('--json', required=True, help='JSON inspection data')
    parser.add_argument('--output', required=True, help='Output Excel path')
    args = parser.parse_args()
    
    try:
        record = json.loads(args.json)
    except json.JSONDecodeError as e:
        print(f'ERROR: Invalid JSON: {e}', file=sys.stderr)
        sys.exit(1)
    
    generate_iqc_excel(record, args.output)


if __name__ == '__main__':
    main()
